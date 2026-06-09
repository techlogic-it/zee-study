import json
import os
import re
import random
from functools import wraps
from flask import (Flask, render_template, request, redirect,
                   url_for, session, flash, jsonify)
from werkzeug.middleware.proxy_fix import ProxyFix
import database as db
import mailer

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)
app.secret_key = os.environ.get('SECRET_KEY', 'gcse-quiz-secret-key-change-in-production')
app.config.update(
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=False,   # Railway terminates TLS at proxy; let browser handle it
)

DIFFICULTY_NAMES = {1: 'Easy', 2: 'Medium', 3: 'Hard', 4: 'Exam Challenge'}
SUBJECTS = ['Biology', 'Chemistry', 'Physics']

TOPICS = {
    'Biology': [
        ('B1', 'Cell Biology'),
        ('B2', 'Organisation'),
        ('B3', 'Infection & Response'),
        ('B4', 'Bioenergetics'),
        ('B5', 'Homeostasis & Response'),
        ('B6', 'Inheritance, Variation & Evolution'),
        ('B7', 'Ecology'),
    ],
    'Chemistry': [
        ('C1', 'Atomic Structure & Periodic Table'),
        ('C2', 'Bonding & Structure'),
        ('C3', 'Quantitative Chemistry'),
        ('C4', 'Chemical Changes'),
        ('C5', 'Energy Changes'),
        ('C6', 'Rate & Extent of Chemical Change'),
        ('C7', 'Organic Chemistry'),
        ('C8', 'Chemical Analysis'),
        ('C9', 'Chemistry of the Atmosphere'),
        ('C10', 'Using Resources'),
    ],
    'Physics': [
        ('P1', 'Energy'),
        ('P2', 'Electricity'),
        ('P3', 'Particle Model of Matter'),
        ('P4', 'Atomic Structure'),
        ('P5', 'Forces'),
        ('P6', 'Waves'),
        ('P7', 'Magnetism & Electromagnetism'),
        ('P8', 'Space Physics'),
    ],
}

def get_topic_name(subject, topic_code):
    for code, name in TOPICS.get(subject, []):
        if code == topic_code:
            return name
    return topic_code


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            # Return 401 JSON for AJAX requests instead of an HTML redirect
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'error': 'Session expired. Please log in again.'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated


@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        first_name = request.form.get('first_name', '').strip()
        last_name  = request.form.get('last_name', '').strip()
        username   = request.form.get('username', '').strip()
        email      = request.form.get('email', '').strip().lower()
        password   = request.form.get('password', '')
        confirm    = request.form.get('confirm_password', '')
        if not first_name or not last_name:
            flash('Please enter your first and last name.', 'error')
        elif not username or not password:
            flash('Username and password are required.', 'error')
        elif password != confirm:
            flash('Passwords do not match.', 'error')
        elif len(password) < 6:
            flash('Password must be at least 6 characters.', 'error')
        elif not email or not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
            flash('Please enter a valid email address.', 'error')
        else:
            ok, msg = db.create_user(username, password, email, first_name, last_name)
            if ok:
                flash('Account created! Please log in.', 'success')
                try:
                    mailer.send_welcome(email, first_name)
                except Exception:
                    pass
                return redirect(url_for('login'))
            flash(msg, 'error')
    return render_template('register.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = db.authenticate_user(username, password)
        if user == 'disabled':
            flash('This account has been disabled. Please contact support.', 'error')
        elif user:
            session['user_id'] = user['user_id']
            session['username'] = user['username']
            return redirect(url_for('dashboard'))
        else:
            flash('Incorrect username or password.', 'error')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    user = db.get_user(session['user_id'])
    recent = db.get_recent_attempts(session['user_id'], limit=5)
    return render_template('dashboard.html', user=user, recent=recent,
                           difficulty_names=DIFFICULTY_NAMES)


@app.route('/topics')
@login_required
def topics():
    return render_template('topics.html')


@app.route('/topic/<subject>')
@login_required
def topic(subject):
    if subject not in SUBJECTS:
        return redirect(url_for('topics'))
    counts = db.get_topic_counts(subject)
    topic_totals = {code: sum(counts.get(code, {}).values()) for code, name in TOPICS[subject]}
    return render_template('topic.html', subject=subject,
                           topics=TOPICS[subject], counts=counts,
                           topic_totals=topic_totals,
                           difficulty_names=DIFFICULTY_NAMES)


@app.route('/difficulty/<subject>')
@login_required
def difficulty_redirect(subject):
    return redirect(url_for('topic', subject=subject))


@app.route('/difficulty/<subject>/<topic_code>')
@login_required
def difficulty(subject, topic_code):
    if subject not in SUBJECTS:
        return redirect(url_for('topics'))
    counts = db.get_topic_counts(subject)
    topic_counts = counts.get(topic_code, {})
    topic_name = get_topic_name(subject, topic_code)
    return render_template('difficulty.html', subject=subject,
                           topic_code=topic_code, topic_name=topic_name,
                           topic_counts=topic_counts,
                           difficulty_names=DIFFICULTY_NAMES)


@app.route('/quiz/<subject>/<topic_code>/<int:level>')
@login_required
def quiz(subject, topic_code, level):
    if subject not in SUBJECTS or level not in DIFFICULTY_NAMES:
        return redirect(url_for('topics'))

    questions = db.get_questions(subject, level, topic_code)
    random.shuffle(questions)
    questions = questions[:10]  # cap at 10
    if not questions:
        flash('No questions available for this selection yet.', 'error')
        return redirect(url_for('difficulty', subject=subject, topic_code=topic_code))

    topic_name = get_topic_name(subject, topic_code)
    questions_display = [
        {
            'question_id': q['question_id'],
            'question_text': q['question_text'],
            'option_a': q['option_a'],
            'option_b': q['option_b'],
            'option_c': q['option_c'],
            'option_d': q['option_d'],
        }
        for q in questions
    ]
    return render_template(
        'quiz.html',
        subject=subject,
        topic_code=topic_code,
        topic_name=topic_name,
        level=level,
        difficulty_name=DIFFICULTY_NAMES[level],
        questions=questions_display,
        questions_json=json.dumps(questions_display),
    )


@app.route('/submit_quiz', methods=['POST'])
@login_required
def submit_quiz():
    import traceback
    try:
        data = request.get_json(force=True, silent=True)
        if not data:
            return jsonify({'error': 'No data received'}), 400

        subject    = data.get('subject')
        level      = int(data.get('level', 1))
        answers    = data.get('answers', {})
        topic_code = data.get('topic_code', None)

        if subject not in SUBJECTS or level not in DIFFICULTY_NAMES:
            return jsonify({'error': 'Invalid subject or level'}), 400

        question_ids = [int(qid) for qid in answers]
        questions    = db.get_questions_by_ids(question_ids)

        score = 0
        wrong_answers = []

        for q in questions:
            qid_str     = str(q['question_id'])
            user_answer = (answers.get(qid_str) or '').upper()
            correct     = q['correct_answer'].upper()
            is_correct  = (user_answer == correct)

            if is_correct:
                score += 1
            else:
                wrong_answers.append({
                    'question_text':      q['question_text'],
                    'user_answer':        user_answer,
                    'user_answer_text':   _option_text(q, user_answer),
                    'correct_answer':     correct,
                    'correct_answer_text': _option_text(q, correct),
                    'explanation':        q['explanation'],
                    'correction_tip':     q['correction_tip'],
                })

        total      = len(questions)
        percentage = round((score / total) * 100) if total else 0

        # XP calculation
        xp = 20
        if percentage >= 70:
            xp += 30
        if percentage == 100:
            xp += 50

        # Streak
        streak_bonus, new_streak = db.update_streak(session['user_id'])
        if streak_bonus:
            xp += 100

        db.add_xp(session['user_id'], xp)
        attempt_id = db.save_attempt(
            session['user_id'], subject, level, score, total, xp, answers, questions, topic_code
        )

    except Exception as e:
        print(f'[submit_quiz] ERROR: {traceback.format_exc()}', flush=True)
        return jsonify({'error': str(e)}), 500

    return jsonify({
        'attempt_id':   attempt_id,
        'score':        score,
        'total':        total,
        'percentage':   percentage,
        'xp_earned':    xp,
        'streak_bonus': streak_bonus,
        'new_streak':   new_streak,
        'wrong_answers': wrong_answers,
        'subject':      subject,
        'difficulty_name': DIFFICULTY_NAMES[level],
    })


def _option_text(question, letter):
    if not letter:
        return '(no answer given)'
    return {
        'A': question['option_a'],
        'B': question['option_b'],
        'C': question['option_c'],
        'D': question['option_d'],
    }.get(letter.upper(), letter)


@app.route('/results/<int:attempt_id>')
@login_required
def results(attempt_id):
    attempt = db.get_attempt(attempt_id, session['user_id'])
    if not attempt:
        return redirect(url_for('dashboard'))
    user = db.get_user(session['user_id'])
    return render_template('results.html', attempt=attempt, user=user,
                           difficulty_names=DIFFICULTY_NAMES)


@app.route('/profile')
@login_required
def profile():
    user    = db.get_user(session['user_id'])
    history = db.get_recent_attempts(session['user_id'], limit=20)
    return render_template('profile.html', user=user, history=history,
                           difficulty_names=DIFFICULTY_NAMES)


@app.route('/account', methods=['GET', 'POST'])
@login_required
def account():
    user = db.get_user(session['user_id'])
    if request.method == 'POST':
        form = request.form.get('form')

        if form == 'details':
            first_name = request.form.get('first_name', '').strip()
            last_name  = request.form.get('last_name', '').strip()
            email      = request.form.get('email', '').strip().lower()
            username   = request.form.get('username', '').strip()
            if not first_name or not last_name or not username:
                flash('Name and username are required.', 'error')
            elif not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
                flash('Please enter a valid email address.', 'error')
            else:
                ok, msg = db.update_user_details(session['user_id'], first_name, last_name, email, username)
                if ok:
                    session['username'] = username  # keep session in sync
                    flash('Your details have been updated.', 'success')
                else:
                    flash(msg, 'error')

        elif form == 'password':
            current  = request.form.get('current_password', '')
            new_pw   = request.form.get('new_password', '')
            confirm  = request.form.get('confirm_password', '')
            verified = db.authenticate_user(user['username'], current)
            if not verified:
                flash('Current password is incorrect.', 'error')
            elif len(new_pw) < 6:
                flash('New password must be at least 6 characters.', 'error')
            elif new_pw != confirm:
                flash('New passwords do not match.', 'error')
            else:
                db.update_user_password(session['user_id'], new_pw)
                flash('Password updated successfully.', 'success')

        return redirect(url_for('account'))
    return render_template('account.html', user=user)


# ── Admin routes ──────────────────────────────────────────
ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'changeme123')
print(f'[admin] Username="{ADMIN_USERNAME}" Password set={"YES (from env)" if os.environ.get("ADMIN_PASSWORD") else "NO (using default: changeme123)"}', flush=True)

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        if username.lower() == ADMIN_USERNAME.lower() and password == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            return redirect(url_for('admin_dashboard'))
        flash('Invalid admin credentials.', 'error')
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_login'))

@app.route('/admin')
@admin_required
def admin_dashboard():
    users = db.get_all_users()
    stats = db.get_admin_stats()
    stats['total_questions'] = db.get_question_count()
    return render_template('admin_dashboard.html', users=users, stats=stats)

@app.route('/admin/user/<int:user_id>/subscription', methods=['POST'])
@admin_required
def admin_update_subscription(user_id):
    plan    = request.form.get('plan', 'free')
    expires = request.form.get('expires') or None
    db.update_user_subscription(user_id, plan, expires)
    flash(f'Subscription updated.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/user/<int:user_id>/edit', methods=['POST'])
@admin_required
def admin_edit_user(user_id):
    first_name = request.form.get('first_name', '').strip()
    last_name  = request.form.get('last_name', '').strip()
    email      = request.form.get('email', '').strip().lower()
    username   = request.form.get('username', '').strip()
    if not first_name or not last_name or not username or not email:
        flash('All fields are required.', 'error')
    else:
        ok, msg = db.admin_update_user(user_id, first_name, last_name, email, username)
        if ok:
            flash('User updated successfully.', 'success')
        else:
            flash(msg, 'error')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/user/<int:user_id>/disable', methods=['POST'])
@admin_required
def admin_disable_user(user_id):
    action = request.form.get('action', 'disable')
    db.set_user_disabled(user_id, action == 'disable')
    flash(f'Account {"disabled" if action == "disable" else "re-enabled"}.', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/user/<int:user_id>/delete', methods=['POST'])
@admin_required
def admin_delete_user(user_id):
    db.delete_user(user_id)
    flash('User account deleted.', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/questions')
@admin_required
def admin_questions():
    subject = request.args.get('subject', '')
    page = int(request.args.get('page', 1))
    questions, total = db.get_all_questions(subject or None, page=page, per_page=25)
    total_pages = max(1, (total + 24) // 25)
    return render_template('admin_questions.html',
                           questions=questions, total=total,
                           page=page, total_pages=total_pages,
                           subject=subject, subjects=SUBJECTS,
                           difficulty_names=DIFFICULTY_NAMES)


@app.route('/admin/question/<int:question_id>/delete', methods=['POST'])
@admin_required
def admin_delete_question(question_id):
    db.delete_question(question_id)
    flash('Question deleted.', 'success')
    subject = request.form.get('subject', '')
    page = request.form.get('page', '1')
    return redirect(url_for('admin_questions', subject=subject, page=page))


@app.route('/admin/upload-questions', methods=['GET', 'POST'])
@admin_required
def admin_upload_questions():
    import io, csv, traceback
    question_count = db.get_question_count()

    if request.method == 'POST':
        if request.form.get('action') == 'delete':
            subject = request.form.get('subject') or None
            deleted = db.delete_questions_by_subject(subject)
            label = subject if subject else 'all subjects'
            flash(f'Deleted {deleted} question(s) from {label}.', 'success')
            return redirect(url_for('admin_upload_questions'))

        file = request.files.get('file')
        if not file or not file.filename:
            flash('No file selected.', 'error')
            return redirect(url_for('admin_upload_questions'))

        filename = file.filename.lower()
        rows = []
        try:
            if filename.endswith('.csv'):
                content = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
            elif filename.endswith('.xlsx'):
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(file.read()))
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(v is not None for v in row):
                        rows.append(dict(zip(headers, row)))
            else:
                flash('Please upload a .csv or .xlsx file.', 'error')
                return redirect(url_for('admin_upload_questions'))

            inserted, skipped = db.insert_questions_bulk(rows)
            flash(f'Done! {inserted} question(s) added, {skipped} row(s) skipped (missing required fields).', 'success')
        except Exception as e:
            print(f'[upload_questions] ERROR: {traceback.format_exc()}', flush=True)
            flash(f'Error reading file: {e}', 'error')

        return redirect(url_for('admin_upload_questions'))

    return render_template('admin_upload_questions.html', question_count=question_count, subjects=SUBJECTS)


@app.route('/admin/question-template')
@admin_required
def admin_question_template():
    from flask import Response
    lines = [
        'subject,topic,topic_code,difficulty,question_text,option_a,option_b,option_c,option_d,correct_answer,explanation,correction_tip',
        'Biology,Cell Biology,B1,1,What is the powerhouse of the cell?,Mitochondria,Nucleus,Ribosome,Cell membrane,A,The mitochondria produce ATP through cellular respiration.,Remember: MITO = energy maker',
        'Chemistry,Atomic Structure & Periodic Table,C1,2,What is the atomic number of Carbon?,6,12,8,14,A,The atomic number equals the number of protons. Carbon has 6 protons.,Atomic number = protons (not mass number).',
    ]
    return Response(
        '\n'.join(lines),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=question_template.csv'}
    )


@app.route('/admin/send-reminders', methods=['POST'])
@admin_required
def admin_send_reminders():
    days = int(request.form.get('days', 7))
    users = db.get_users_for_reminder(days_inactive=days)
    sent = 0
    for u in users:
        if u.get('email'):
            mailer.send_reminder(u['email'], u['username'], days)
            sent += 1
    flash(f'Reminder emails sent to {sent} user(s) inactive for {days}+ days.', 'success')
    return redirect(url_for('admin_dashboard'))


if __name__ == '__main__':
    db.init_db()
    from seed_questions import seed
    if not db.questions_exist():
        seed()
    app.run(debug=True, port=5001)
